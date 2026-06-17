#!/usr/bin/env python3
"""周末返校人数统计系统"""

import json
import os
from datetime import datetime, timedelta
from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# ============ 班级数据（去掉23级和24级3+2） ============
CLASSES = [
    {"id": 1,  "name": "24级会计2班",      "major": "大数据与会计",       "count": 44, "counselor": "马小丽"},
    {"id": 2,  "name": "24级会计3班",      "major": "大数据与会计",       "count": 43, "counselor": "杨阳"},
    {"id": 3,  "name": "24级会计4班",      "major": "大数据与会计",       "count": 42, "counselor": "杨阳"},
    {"id": 4,  "name": "24级会管1班",      "major": "会计信息管理",       "count": 44, "counselor": "柴泳泳"},
    {"id": 5,  "name": "24级财务1班",      "major": "大数据与财务管理",   "count": 43, "counselor": "张敏"},
    {"id": 6,  "name": "24级财务2班",      "major": "大数据与财务管理",   "count": 41, "counselor": "崔淑林"},
    {"id": 7,  "name": "24级电商1班",      "major": "电子商务",           "count": 40, "counselor": "郗轶君"},
    {"id": 8,  "name": "24级电商2班",      "major": "电子商务",           "count": 41, "counselor": "郗轶君"},
    {"id": 9,  "name": "24级市营1班",      "major": "市场营销",           "count": 35, "counselor": "柴泳泳"},
    {"id": 10, "name": "24级物流1班",      "major": "现代物流管理",       "count": 46, "counselor": "柴泳泳"},
    {"id": 11, "name": "25级会计一班",     "major": "大数据与会计",       "count": 40, "counselor": "耿文"},
    {"id": 12, "name": "25级会计二班",     "major": "大数据与会计",       "count": 42, "counselor": "耿文"},
    {"id": 13, "name": "25级会计三班",     "major": "大数据与会计",       "count": 42, "counselor": "李知之"},
    {"id": 14, "name": "25级会计四班",     "major": "大数据与会计",       "count": 44, "counselor": "李知之"},
    {"id": 15, "name": "25级财务一班",     "major": "大数据与财务管理",   "count": 40, "counselor": "王强强"},
    {"id": 16, "name": "25级财务二班",     "major": "大数据与财务管理",   "count": 38, "counselor": "王强强"},
    {"id": 17, "name": "25级电商一班",     "major": "电子商务",           "count": 46, "counselor": "郗轶君"},
    {"id": 18, "name": "25级物流一班",     "major": "现代物流管理",       "count": 46, "counselor": "张嘉愉"},
]

COUNSELORS = sorted(set(c["counselor"] for c in CLASSES))

# ============ 数据存储 ============
def get_recent_dates():
    """获取最近若干天的日期列表，从今天开始"""
    today = datetime.now()
    dates = []
    for i in range(14):
        d = today - timedelta(days=i)
        dates.append(d.strftime("%Y-%m-%d"))
    dates.reverse()
    return dates

def data_file(week_date):
    return os.path.join(DATA_DIR, f"week_{week_date}.json")

def load_week_data(week_date):
    fp = data_file(week_date)
    if os.path.exists(fp):
        with open(fp, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_week_data(week_date, data):
    fp = data_file(week_date)
    with open(fp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ============ 页面模板 ============
COMMON_HEAD = '''
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>财贸系周末返校人数统计</title>
<style>
  :root { --primary: #4F46E5; --primary-light: #6366F1; --bg: #F9FAFB; --card: #FFFFFF; --text: #111827; --muted: #6B7280; --border: #E5E7EB; --success: #10B981; --warning: #F59E0B; --danger: #EF4444; }
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: var(--bg); color: var(--text); line-height: 1.6; }
  .container { max-width: 960px; margin: 0 auto; padding: 20px; }
  .header { text-align: center; padding: 30px 0 20px; }
  .header h1 { font-size: 1.6rem; color: var(--primary); margin-bottom: 6px; }
  .header p { color: var(--muted); font-size: 0.95rem; }
  .card { background: var(--card); border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,.08); }
  .btn { display: inline-block; padding: 10px 24px; border-radius: 8px; border: none; font-size: .95rem; font-weight: 600; cursor: pointer; transition: all .2s; text-decoration: none; }
  .btn-primary { background: var(--primary); color: #fff; }
  .btn-primary:hover { background: var(--primary-light); }
  .btn-success { background: var(--success); color: #fff; }
  .btn-success:hover { opacity: .9; }
  .btn-outline { background: transparent; color: var(--primary); border: 1.5px solid var(--primary); }
  .btn-outline:hover { background: var(--primary); color: #fff; }
  table { width: 100%; border-collapse: collapse; font-size: .88rem; }
  th { background: #F3F4F6; font-weight: 600; text-align: center; padding: 10px 6px; border-bottom: 2px solid var(--border); white-space: nowrap; }
  td { padding: 9px 6px; text-align: center; border-bottom: 1px solid var(--border); vertical-align: middle; }
  tr:hover td { background: #F9FAFB; }
  input[type="number"] { width: 68px; padding: 6px 8px; border: 1.5px solid var(--border); border-radius: 6px; text-align: center; font-size: .9rem; transition: border-color .2s; }
  input[type="number"]:focus { outline: none; border-color: var(--primary); box-shadow: 0 0 0 3px rgba(79,70,229,.1); }
  input[type="text"] { width: 100%; min-width: 120px; padding: 6px 10px; border: 1.5px solid var(--border); border-radius: 6px; font-size: .85rem; transition: border-color .2s; }
  input[type="text"]:focus { outline: none; border-color: var(--primary); box-shadow: 0 0 0 3px rgba(79,70,229,.1); }
  select { padding: 8px 14px; border: 1.5px solid var(--border); border-radius: 8px; font-size: .95rem; background: #fff; cursor: pointer; }
  select:focus { outline: none; border-color: var(--primary); }
  .badge { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: .78rem; font-weight: 600; }
  .badge-green { background: #D1FAE5; color: #065F46; }
  .badge-yellow { background: #FEF3C7; color: #92400E; }
  .badge-gray { background: #F3F4F6; color: #6B7280; }
  .progress-bar { width: 100%; height: 6px; background: #E5E7EB; border-radius: 3px; overflow: hidden; margin-top: 6px; }
  .progress-fill { height: 100%; border-radius: 3px; transition: width .4s; }
  .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); gap: 12px; margin-bottom: 20px; }
  .stat-card { background: var(--card); border-radius: 10px; padding: 16px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,.06); }
  .stat-card .num { font-size: 1.8rem; font-weight: 700; }
  .stat-card .label { font-size: .82rem; color: var(--muted); margin-top: 2px; }
  .toast { position: fixed; top: 20px; right: 20px; padding: 14px 24px; border-radius: 10px; color: #fff; font-weight: 600; z-index: 999; transform: translateX(120%); transition: transform .3s; }
  .toast.show { transform: translateX(0); }
  .toast-success { background: var(--success); }
  .empty-state { text-align: center; padding: 40px 20px; color: var(--muted); }
  .empty-state svg { width: 60px; height: 60px; margin-bottom: 12px; opacity: .4; }
  .hint { font-size: .8rem; color: var(--muted); margin-top: 4px; }
  .table-wrap { overflow-x: auto; }
  @media (max-width: 640px) {
    .container { padding: 12px; }
    .header h1 { font-size: 1.3rem; }
    th, td { padding: 7px 4px; font-size: .8rem; }
    input[type="number"] { width: 56px; }
    input[type="text"] { min-width: 90px; font-size: .8rem; }
  }
</style>
'''

INDEX_PAGE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>''' + COMMON_HEAD + '''</head>
<body>
<div class="container">
  <div class="header">
    <h1>📊 财贸系周末返校人数统计</h1>
    <p>请选择统计日期</p>
  </div>
  <div class="card">
    <label for="week" style="font-weight:600;margin-right:10px;">统计日期：</label>
    <select id="week" onchange="updateLinks()"></select>
  </div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px;">
    <a id="fillLink" class="card" style="display:block;text-decoration:none;text-align:center;padding:32px 20px;cursor:pointer;transition:transform .2s;" onmouseover="this.style.transform='translateY(-2px)'" onmouseout="this.style.transform=''">
      <div style="font-size:2rem;margin-bottom:8px;">✏️</div>
      <div style="font-weight:700;font-size:1.1rem;color:var(--primary);">辅导员填写入口</div>
      <div style="color:var(--muted);font-size:.85rem;margin-top:4px;">选择辅导员后填写返校人数</div>
    </a>
    <a id="adminLink" class="card" style="display:block;text-decoration:none;text-align:center;padding:32px 20px;cursor:pointer;transition:transform .2s;" onmouseover="this.style.transform='translateY(-2px)'" onmouseout="this.style.transform=''">
      <div style="font-size:2rem;margin-bottom:8px;">📋</div>
      <div style="font-weight:700;font-size:1.1rem;color:var(--primary);">管理员查看 / 导出</div>
      <div style="color:var(--muted);font-size:.85rem;margin-top:4px;">查看汇总 & 下载Excel</div>
    </a>
  </div>
  <div class="card" style="text-align:center;color:var(--muted);font-size:.85rem;">
    共 <strong>''' + str(len(CLASSES)) + '''</strong> 个班级 · <strong>''' + str(len(COUNSELORS)) + '''</strong> 位辅导员
  </div>
</div>
<script>
const fridays = ''' + json.dumps(get_recent_dates(), ensure_ascii=False) + ''';
const sel = document.getElementById('week');
fridays.forEach((f, i) => {
  const opt = document.createElement('option');
  opt.value = f;
  const d = new Date(f);
  const weekDays = ['周日','周一','周二','周三','周四','周五','周六'];
  opt.textContent = (d.getMonth()+1) + '月' + d.getDate() + '日（' + weekDays[d.getDay()] + '）';
  // 默认选中今天
  const today = new Date();
  if (d.getFullYear()===today.getFullYear() && d.getMonth()===today.getMonth() && d.getDate()===today.getDate()) {
    opt.selected = true;
  }
  sel.appendChild(opt);
});
function updateLinks() {
  const w = sel.value;
  document.getElementById('fillLink').href = '/fill?week=' + w;
  document.getElementById('adminLink').href = '/admin?week=' + w;
}
updateLinks();
</script>
</body>
</html>
'''

FILL_PAGE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>''' + COMMON_HEAD + '''</head>
<body>
<div id="toast" class="toast toast-success">保存成功！</div>
<div class="container">
  <div class="header">
    <h1>✏️ 周末返校人数填写</h1>
    <p id="weekLabel"></p>
  </div>
  <div class="card">
    <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
      <label for="counselor" style="font-weight:600;">选择辅导员：</label>
      <select id="counselor" onchange="loadData()">
        <option value="">-- 请选择 --</option>
        ''' + ''.join(f'<option value="{c}">{c}</option>' for c in COUNSELORS) + '''
      </select>
    </div>
  </div>
  <div id="formArea" style="display:none;">
    <div class="card" style="margin-bottom:8px;color:var(--muted);font-size:.85rem;">
      💡 <strong>备注</strong>栏请填写<strong>未按时返校</strong>的学生姓名，多人用顿号分隔
    </div>
    <div class="card">
      <div class="table-wrap">
      <table>
        <thead>
          <tr><th>序号</th><th>班级</th><th>专业</th><th>班级人数</th><th>返校人数</th><th>未返校</th><th>备注（未返校学生姓名）</th></tr>
        </thead>
        <tbody id="tbody"></tbody>
        <tfoot>
          <tr style="font-weight:700;background:#F9FAFB;">
            <td colspan="3">合计</td>
            <td id="totalAll">0</td>
            <td id="totalReturn">0</td>
            <td id="totalAbsent">0</td>
            <td></td>
          </tr>
        </tfoot>
      </table>
      </div>
    </div>
    <div style="text-align:center;padding:10px 0 30px;">
      <button class="btn btn-primary" onclick="saveData()" style="margin-right:12px;">💾 提交保存</button>
      <a class="btn btn-outline" href="/">返回首页</a>
    </div>
  </div>
  <div id="emptyHint" class="card empty-state" style="display:none;">
    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5"><path d="M15.75 6a3.75 3.75 0 11-7.5 0 3.75 3.75 0 017.5 0zM4.501 20.118a7.5 7.5 0 0114.998 0"/></svg>
    <p>请先选择辅导员姓名</p>
  </div>
</div>
<script>
const week = new URLSearchParams(location.search).get('week');
const d = new Date(week);
document.getElementById('weekLabel').textContent = '统计日期：' + (d.getMonth()+1) + '月' + d.getDate() + '日';

const allClasses = ''' + json.dumps(CLASSES, ensure_ascii=False) + ''';
const counselorsList = ''' + json.dumps(COUNSELORS, ensure_ascii=False) + ''';

function loadData() {
  const name = document.getElementById('counselor').value;
  const formArea = document.getElementById('formArea');
  const emptyHint = document.getElementById('emptyHint');
  if (!name) { formArea.style.display='none'; emptyHint.style.display='block'; return; }
  formArea.style.display='block'; emptyHint.style.display='none';

  const myClasses = allClasses.filter(c => c.counselor === name);
  fetch('/api/data?week=' + week + '&counselor=' + encodeURIComponent(name))
    .then(r => r.json())
    .then(saved => {
      const tbody = document.getElementById('tbody');
      tbody.innerHTML = '';
      myClasses.forEach((cls, i) => {
        const classData = saved[cls.name] || {};
        const retVal = classData.return !== undefined ? classData.return : '';
        const remarkVal = classData.remark || '';
        const absent = retVal !== '' ? cls.count - retVal : '';
        const tr = document.createElement('tr');
        tr.innerHTML = `
          <td>${i+1}</td>
          <td style="font-weight:600;white-space:nowrap;">${cls.name}</td>
          <td style="white-space:nowrap;">${cls.major}</td>
          <td>${cls.count}</td>
          <td><input type="number" min="0" max="${cls.count}" value="${retVal}" data-class="${cls.name}" data-total="${cls.count}" oninput="calcRow(this)"></td>
          <td class="absent-cell" style="font-weight:600;color:${retVal!==''?(cls.count-retVal>0?'var(--danger)':'var(--success)'):'var(--muted)'}">${retVal!==''?absent:'-'}</td>
          <td><input type="text" data-class="${cls.name}" placeholder="如：张三、李四" value="${remarkVal}"></td>
        `;
        tbody.appendChild(tr);
      });
      calcTotal();
    });
}

function calcRow(input) {
  const total = parseInt(input.dataset.total);
  let val = input.value === '' ? '' : Math.min(Math.max(parseInt(input.value)||0, 0), total);
  if (input.value !== '') input.value = val;
  const tr = input.closest('tr');
  const absentTd = tr.querySelector('.absent-cell');
  if (val === '') { absentTd.textContent = '-'; absentTd.style.color = 'var(--muted)'; }
  else { const absent = total - val; absentTd.textContent = absent; absentTd.style.color = absent > 0 ? 'var(--danger)' : 'var(--success)'; }
  calcTotal();
}

function calcTotal() {
  const name = document.getElementById('counselor').value;
  let tAll=0, tRet=0;
  allClasses.filter(c => c.counselor === name).forEach(c => tAll+=c.count);
  document.querySelectorAll('#tbody input[type=number]').forEach(inp => {
    if(inp.value!=='') tRet+=parseInt(inp.value);
  });
  document.getElementById('totalAll').textContent = tAll;
  document.getElementById('totalReturn').textContent = tRet;
  document.getElementById('totalAbsent').textContent = tAll - tRet;
}

function saveData() {
  const name = document.getElementById('counselor').value;
  if (!name) return;
  const data = {};
  const rows = document.querySelectorAll('#tbody tr');
  rows.forEach(tr => {
    const numInput = tr.querySelector('input[type=number]');
    const textInput = tr.querySelector('input[type=text]');
    if (!numInput) return;
    const className = numInput.dataset.class;
    const retVal = numInput.value !== '' ? parseInt(numInput.value) : null;
    const remark = textInput ? textInput.value.trim() : '';
    if (retVal !== null) {
      data[className] = { return: retVal, remark: remark };
    }
  });
  fetch('/api/data?week=' + week + '&counselor=' + encodeURIComponent(name), {
    method: 'POST', headers: {'Content-Type':'application/json'}, body: JSON.stringify(data)
  }).then(r => r.json()).then(res => {
    if (res.ok) {
      const toast = document.getElementById('toast');
      toast.classList.add('show');
      setTimeout(() => toast.classList.remove('show'), 2000);
    }
  });
}

// auto-load if counselor in URL hash
const hashCounselor = decodeURIComponent(location.hash.slice(1));
if (hashCounselor && counselorsList.includes(hashCounselor)) {
  document.getElementById('counselor').value = hashCounselor;
  loadData();
}
</script>
</body>
</html>
'''

ADMIN_PAGE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>''' + COMMON_HEAD + '''</head>
<body>
<div class="container">
  <div class="header">
    <h1>📋 返校统计汇总</h1>
    <p id="weekLabel"></p>
  </div>
  <div class="stats-grid" id="statsGrid"></div>
  <div class="card" style="margin-bottom:16px;">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;margin-bottom:16px;">
      <h3 style="font-size:1rem;">各辅导员填写进度</h3>
      <div>
        <a class="btn btn-success" id="downloadBtn" href="#" style="margin-right:8px;">📥 下载Excel</a>
        <a class="btn btn-outline" href="/">返回首页</a>
      </div>
    </div>
    <div id="progressArea"></div>
  </div>
  <div class="card">
    <h3 style="font-size:1rem;margin-bottom:12px;">详细数据</h3>
    <div class="table-wrap">
      <table>
        <thead>
          <tr><th>序号</th><th>班级</th><th>专业</th><th>辅导员</th><th>班级人数</th><th>返校人数</th><th>未返校</th><th>返校率</th><th>备注（未返校学生）</th></tr>
        </thead>
        <tbody id="detailBody"></tbody>
        <tfoot>
          <tr style="font-weight:700;background:#F3F4F6;">
            <td colspan="4">合计</td>
            <td id="sumTotal">0</td>
            <td id="sumReturn">0</td>
            <td id="sumAbsent">0</td>
            <td id="sumRate">-</td>
            <td></td>
          </tr>
        </tfoot>
      </table>
    </div>
  </div>
</div>
<script>
const week = new URLSearchParams(location.search).get('week');
const d = new Date(week);
document.getElementById('weekLabel').textContent = '统计日期：' + (d.getMonth()+1) + '月' + d.getDate() + '日';
document.getElementById('downloadBtn').href = '/api/export?week=' + week;

const allClasses = ''' + json.dumps(CLASSES, ensure_ascii=False) + ''';
const counselors = ''' + json.dumps(COUNSELORS, ensure_ascii=False) + ''';

fetch('/api/alldata?week=' + week).then(r=>r.json()).then(allData => {
  let totalStudents = 0, totalReturn = 0, filledCounselors = 0;
  allClasses.forEach(c => {
    totalStudents += c.count;
    const saved = allData[c.counselor];
    if (saved && saved[c.name] && saved[c.name].return !== undefined) totalReturn += saved[c.name].return;
  });
  counselors.forEach(c => {
    const s = allData[c] || {};
    const hasData = allClasses.filter(cl => cl.counselor === c).some(cl => s[cl.name] && s[cl.name].return !== undefined);
    if (hasData) filledCounselors++;
  });
  const returnRate = totalStudents > 0 ? (totalReturn/totalStudents*100).toFixed(1) : '-';

  document.getElementById('statsGrid').innerHTML = `
    <div class="stat-card"><div class="num" style="color:var(--primary)">${totalStudents}</div><div class="label">学生总人数</div></div>
    <div class="stat-card"><div class="num" style="color:var(--success)">${totalReturn}</div><div class="label">已返校人数</div></div>
    <div class="stat-card"><div class="num" style="color:var(--danger)">${totalStudents - totalReturn}</div><div class="label">未返校人数</div></div>
    <div class="stat-card"><div class="num" style="color:var(--warning)">${returnRate}%</div><div class="label">返校率</div></div>
    <div class="stat-card"><div class="num" style="color:var(--primary)">${filledCounselors}/${counselors.length}</div><div class="label">辅导员填写</div></div>
  `;

  // progress
  const progArea = document.getElementById('progressArea');
  progArea.innerHTML = '';
  counselors.forEach(c => {
    const myClasses = allClasses.filter(cl => cl.counselor === c);
    const saved = allData[c] || {};
    const filled = myClasses.filter(cl => saved[cl.name] && saved[cl.name].return !== undefined).length;
    const pct = Math.round(filled / myClasses.length * 100);
    const badge = pct === 100 ? '<span class="badge badge-green">已完成</span>' : pct > 0 ? '<span class="badge badge-yellow">填写中</span>' : '<span class="badge badge-gray">未填写</span>';
    progArea.innerHTML += `
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;">
        <span style="min-width:70px;font-weight:600;">${c}</span>
        ${badge}
        <div class="progress-bar" style="flex:1;"><div class="progress-fill" style="width:${pct}%;background:${pct===100?'var(--success)':pct>0?'var(--warning)':'#D1D5DB'};"></div></div>
        <span style="font-size:.82rem;color:var(--muted);min-width:50px;text-align:right;">${filled}/${myClasses.length}</span>
      </div>
    `;
  });

  // detail table
  const tbody = document.getElementById('detailBody');
  tbody.innerHTML = '';
  let sTotal=0, sReturn=0;
  allClasses.forEach((cls, i) => {
    const saved = allData[cls.counselor] || {};
    const classData = saved[cls.name] || {};
    const ret = classData.return !== undefined ? classData.return : null;
    const remark = classData.remark || '';
    const absent = ret !== null ? cls.count - ret : null;
    const rate = ret !== null ? (ret/cls.count*100).toFixed(1)+'%' : '-';
    sTotal += cls.count;
    sReturn += (ret || 0);
    const row = document.createElement('tr');
    row.innerHTML = `
      <td>${i+1}</td>
      <td style="font-weight:600;white-space:nowrap;">${cls.name}</td>
      <td style="white-space:nowrap;">${cls.major}</td>
      <td>${cls.counselor}</td>
      <td>${cls.count}</td>
      <td style="color:${ret!==null?'var(--success)':'var(--muted)'};font-weight:${ret!==null?700:400};">${ret!==null?ret:'未填写'}</td>
      <td style="color:${absent!==null&&absent>0?'var(--danger)':'var(--muted)'};">${absent!==null?absent:'-'}</td>
      <td>${rate}</td>
      <td style="text-align:left;font-size:.82rem;color:${remark?'var(--danger)':'var(--muted)'};max-width:200px;word-break:break-all;">${remark||'-'}</td>
    `;
    tbody.appendChild(row);
  });
  document.getElementById('sumTotal').textContent = sTotal;
  document.getElementById('sumReturn').textContent = sReturn;
  document.getElementById('sumAbsent').textContent = sTotal - sReturn;
  document.getElementById('sumRate').textContent = sTotal > 0 ? (sReturn/sTotal*100).toFixed(1)+'%' : '-';
});
</script>
</body>
</html>
'''

# ============ 路由 ============
@app.route('/')
def index():
    return render_template_string(INDEX_PAGE)

@app.route('/fill')
def fill_page():
    return render_template_string(FILL_PAGE)

@app.route('/admin')
def admin_page():
    return render_template_string(ADMIN_PAGE)

@app.route('/api/data', methods=['GET', 'POST'])
def api_data():
    week = request.args.get('week')
    counselor = request.args.get('counselor')
    if not week or not counselor:
        return jsonify({})

    all_data = load_week_data(week)
    if counselor not in all_data:
        all_data[counselor] = {}

    if request.method == 'GET':
        return jsonify(all_data.get(counselor, {}))
    else:
        submitted = request.get_json(force=True)
        all_data[counselor] = submitted
        save_week_data(week, all_data)
        return jsonify({"ok": True})

@app.route('/api/alldata')
def api_alldata():
    week = request.args.get('week')
    if not week:
        return jsonify({})
    return jsonify(load_week_data(week))

@app.route('/api/export')
def api_export():
    week = request.args.get('week')
    if not week:
        return "缺少周次参数", 400

    all_data = load_week_data(week)
    d = datetime.strptime(week, "%Y-%m-%d")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "返校统计"

    # styles
    title_font = Font(name='微软雅黑', bold=True, size=14)
    normal_font = Font(name='微软雅黑', size=10)
    bold_font = Font(name='微软雅黑', bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')

    # title
    ws.merge_cells('A1:I1')
    ws['A1'] = f'财贸系周末返校人数统计（{d.month}月{d.day}日）'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    # headers
    headers = ['序号', '班级', '专业', '辅导员', '班级人数', '返校人数', '未返校人数', '返校率', '备注（未返校学生姓名）']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # data
    for i, cls in enumerate(CLASSES):
        row = i + 3
        saved = all_data.get(cls['counselor'], {})
        raw = saved.get(cls['name'])
        # 兼容旧格式（纯数字）和新格式（{return, remark}）
        if isinstance(raw, dict):
            ret = raw.get('return')
            remark = raw.get('remark', '')
        elif isinstance(raw, int):
            ret = raw
            remark = ''
        else:
            ret = None
            remark = ''
        absent = cls['count'] - ret if ret is not None else None
        rate = f'{ret/cls["count"]*100:.1f}%' if ret is not None else ''

        vals = [i+1, cls['name'], cls['major'], cls['counselor'], cls['count'],
                ret if ret is not None else '', absent if absent is not None else '', rate, remark]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = normal_font
            cell.alignment = left_wrap if col == 9 else center
            cell.border = thin_border

    # total row
    total_row = len(CLASSES) + 3
    total_students = sum(c['count'] for c in CLASSES)
    def _get_return(class_raw):
        """兼容旧格式和新格式，提取返校人数"""
        if isinstance(class_raw, dict):
            return class_raw.get('return')
        elif isinstance(class_raw, int):
            return class_raw
        return None

    total_return = sum(
        _get_return(all_data.get(c['counselor'], {}).get(c['name'])) or 0
        for c in CLASSES
        if _get_return(all_data.get(c['counselor'], {}).get(c['name'])) is not None
    )
    total_absent = total_students - total_return
    total_rate = f'{total_return/total_students*100:.1f}%' if total_students > 0 else ''

    ws.merge_cells(f'A{total_row}:D{total_row}')
    total_vals = ['合计', total_students, total_return, total_absent, total_rate, '']
    total_cols = [1, 5, 6, 7, 8, 9]
    for v, col in zip(total_vals, total_cols):
        cell = ws.cell(row=total_row, column=col, value=v)
        cell.font = bold_font
        cell.alignment = center
        cell.border = thin_border
    for col in range(2, 5):
        ws.cell(row=total_row, column=col).border = thin_border

    # column widths
    widths = [6, 20, 18, 10, 10, 10, 10, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # counselor summary sheet
    ws2 = wb.create_sheet("辅导员汇总")
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f'辅导员填写进度（{d.month}月{d.day}日）'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center

    sum_headers = ['辅导员', '负责班级数', '已填写班级数', '返校人数', '返校率', '未返校学生']
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for i, c in enumerate(COUNSELORS):
        row = i + 3
        my_classes = [cl for cl in CLASSES if cl['counselor'] == c]
        saved = all_data.get(c, {})
        filled = sum(1 for cl in my_classes if _get_return(saved.get(cl['name'])) is not None)
        ret_count = sum(_get_return(saved.get(cl['name'])) or 0 for cl in my_classes if _get_return(saved.get(cl['name'])) is not None)
        total_c = sum(cl['count'] for cl in my_classes)
        rate_c = f'{ret_count/total_c*100:.1f}%' if total_c > 0 else ''
        # collect all absent names
        absent_names = []
        for cl in my_classes:
            raw = saved.get(cl['name'], {})
            if isinstance(raw, dict) and raw.get('remark'):
                absent_names.append(raw['remark'])
        absent_str = '；'.join(absent_names) if absent_names else ''
        vals = [c, len(my_classes), filled, ret_count, rate_c, absent_str]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.font = normal_font
            cell.alignment = left_wrap if col == 6 else center
            cell.border = thin_border

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 40

    filename = f'财贸系返校情况{d.month}月{d.day}日.xlsx'
    filepath = os.path.join(DATA_DIR, f'export_{week}.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
